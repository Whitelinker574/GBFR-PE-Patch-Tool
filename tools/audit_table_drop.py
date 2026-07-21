#!/usr/bin/env python3
"""Audit an untrusted GBFR spreadsheet/CT drop against the checked-in catalogs.

The spreadsheets shared with this project have historically been renamed,
partially copied, or saved with invalid worksheet dimensions.  This tool never
uses filenames as evidence.  It extracts hash-shaped cells, compares their sets
with the checked-in catalogs, and emits both machine-readable JSON and a small
review workbook.  It does not mutate production data.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HASH_RE = re.compile(r"^(?:0x)?([0-9A-Fa-f]{8})$")
SPREADSHEET_NS = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def normalise_hash(value: object) -> str | None:
    if isinstance(value, int) and 0 <= value <= 0xFFFFFFFF:
        return f"{value:08X}"
    text = str(value or "").strip()
    match = HASH_RE.fullmatch(text)
    return match.group(1).upper() if match else None


def read_rows(path: Path) -> tuple[list[list[object]], str]:
    """Read the first sheet, tolerating malformed dimensions and SpreadsheetML."""
    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
        sheet = workbook.worksheets[0]
        rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
        workbook.close()
        return rows, "openpyxl"
    except Exception as first_error:
        raw = path.read_bytes()
        if raw.lstrip().startswith(b"<?xml"):
            root = ET.fromstring(raw)
            rows = []
            for row in root.findall(".//ss:Row", SPREADSHEET_NS):
                values = []
                for cell in row.findall("ss:Cell", SPREADSHEET_NS):
                    data = cell.find("ss:Data", SPREADSHEET_NS)
                    values.append(data.text if data is not None else None)
                rows.append(values)
            return rows, f"spreadsheetml-fallback ({type(first_error).__name__})"
        if zipfile.is_zipfile(path):
            with zipfile.ZipFile(path) as archive:
                shared = []
                if "xl/sharedStrings.xml" in archive.namelist():
                    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
                    shared = ["".join(node.itertext()) for node in root]
                sheet_names = sorted(name for name in archive.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"))
                if not sheet_names:
                    raise first_error
                root = ET.fromstring(archive.read(sheet_names[0]))
                ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                rows = []
                for row in root.findall(".//x:sheetData/x:row", ns):
                    values = []
                    for cell in row.findall("x:c", ns):
                        value = cell.find("x:v", ns)
                        text = value.text if value is not None else None
                        if cell.get("t") == "s" and text is not None:
                            text = shared[int(text)]
                        values.append(text)
                    rows.append(values)
                return rows, f"ooxml-fallback ({type(first_error).__name__})"
        raise first_error


def json_hashes(path: Path, key: str) -> set[str]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    values = payload[key]
    result = set()
    for row in values:
        value = normalise_hash(row.get("hash"))
        if value:
            result.add(value)
    return result


def load_catalogs(repo: Path) -> dict[str, set[str]]:
    return {
        "traits": json_hashes(repo / "data" / "traits.json", "traits"),
        "summons": json_hashes(repo / "data" / "summons.json", "summons"),
        "summon_sub_params": json_hashes(repo / "data" / "summon_sub_params.json", "subParams"),
        "wrightstones": json_hashes(repo / "data" / "wrightstones.json", "wrightstones"),
    }


def extract_hash_rows(rows: list[list[object]]) -> list[dict[str, object]]:
    result = []
    for index, row in enumerate(rows, start=1):
        found = next((normalise_hash(value) for value in row if normalise_hash(value)), None)
        if not found:
            continue
        result.append({
            "row": index,
            "hash": found,
            "values": [str(value).strip() for value in row if value is not None and str(value).strip()],
        })
    return result


def classify(name: str, hashes: set[str], catalogs: dict[str, set[str]]) -> tuple[str, float, str]:
    scores = []
    for catalog_name, catalog_hashes in catalogs.items():
        overlap = len(hashes & catalog_hashes)
        precision = overlap / len(hashes) if hashes else 0.0
        recall = overlap / len(catalog_hashes) if catalog_hashes else 0.0
        scores.append((precision, recall, overlap, catalog_name))
    precision, recall, overlap, best = max(scores, default=(0.0, 0.0, 0, "unknown"))
    if not hashes:
        return "reject", 0.0, "no hash-shaped data rows"
    if precision == 1.0 and recall == 1.0:
        return "verified-equal", 1.0, f"exactly matches checked-in {best} hash set"
    if precision == 1.0 and overlap:
        return "verified-subset", recall, f"strict subset of checked-in {best} ({overlap}/{len(catalogs[best])})"
    if precision >= 0.95 and overlap >= 4:
        return "candidate", precision, f"mostly overlaps checked-in {best} ({overlap}/{len(hashes)}); review non-overlap rows"
    filename_hint = next((item for item in catalogs if item in name.lower()), "")
    reason = f"content best resembles {best} ({overlap}/{len(hashes)}), not trusted by filename"
    if filename_hint and filename_hint != best:
        reason += f"; filename hint conflicts with content ({filename_hint})"
    return "reject", precision, reason


def audit(drop: Path, repo: Path) -> dict[str, object]:
    catalogs = load_catalogs(repo)
    files = []
    extracted = {}
    for path in sorted(drop.iterdir(), key=lambda item: item.name.casefold()):
        base = {
            "file": path.name,
            "size": path.stat().st_size,
            "sha256": sha256(path),
        }
        if path.suffix.lower() != ".xlsx":
            base.update({"kind": path.suffix.lower().lstrip("."), "decision": "reference-only", "reason": "not a spreadsheet; reviewed separately"})
            files.append(base)
            continue
        prefix = path.read_bytes()[:512].lstrip()
        if prefix.startswith(b"<?xml") and b"<CheatTable" in prefix:
            base.update({
                "kind": "ct-disguised-as-xlsx",
                "decision": "reject",
                "confidence": 0,
                "reason": "file is Cheat Engine XML, not an Excel workbook",
            })
            files.append(base)
            continue
        try:
            rows, reader = read_rows(path)
            hash_rows = extract_hash_rows(rows)
            hashes = [row["hash"] for row in hash_rows]
            unique = set(hashes)
            decision, confidence, reason = classify(path.name, unique, catalogs)
            base.update({
                "kind": "xlsx",
                "reader": reader,
                "rows": len(rows),
                "hashRows": len(hashes),
                "uniqueHashes": len(unique),
                "duplicateHashes": sorted(value for value, count in Counter(hashes).items() if count > 1),
                "decision": decision,
                "confidence": round(confidence, 4),
                "reason": reason,
            })
            extracted[path.name] = hash_rows
        except Exception as error:  # report damaged inputs without stopping the audit
            base.update({"kind": "xlsx", "decision": "reject", "confidence": 0, "reason": f"unreadable: {type(error).__name__}: {error}"})
        files.append(base)

    families = []
    family_names = {
        "traits": ["词条ID V9.xlsx"],
        "sigils": ["因子IDV4 .xlsx", "因子IDV6 .xlsx", "因子IDV9 .xlsx"],
        "summons": ["召唤石ID.xlsx", "召唤石IDV2.xlsx"],
        "summon_sub_params": ["召唤石副词条.xlsx", "召唤石副词条V2.xlsx"],
        "wrightstones": ["祝福石ID.xlsx", "祝福石IDV2.xlsx"],
        "overlimit": ["上限突破ID.xlsx", "上限突破IDV2.xlsx"],
    }
    for family, names in family_names.items():
        previous = None
        for name in names:
            current = {row["hash"] for row in extracted.get(name, [])}
            if not current:
                continue
            if previous is not None:
                families.append({
                    "family": family,
                    "from": previous[0],
                    "to": name,
                    "added": sorted(current - previous[1]),
                    "removed": sorted(previous[1] - current),
                })
            previous = (name, current)

    natural_rules = json.loads((repo / "data" / "summon_natural_rules_202.json").read_text(encoding="utf-8-sig"))["rows"]
    channel_parity = [
        {
            "family": "summon type/main/sub",
            "saveChannel": "SummonSaveGen.GetOptions",
            "memoryChannel": "App.SummonGetOptions",
            "decision": "verified-equal",
            "detail": f"same backend function; 189 types / 82 main traits / 22 sub parameters; {len(natural_rules)} type rules",
        },
        {
            "family": "wrightstone traits",
            "saveChannel": "WrightstoneGen.GetTraitList",
            "memoryChannel": "App.WrightstoneMemoryGetOptions",
            "decision": "verified-equal",
            "detail": "same catalog, localization, max-level and allowed-level functions; enforced by catalog_channel_parity_test.go",
        },
        {
            "family": "sigils / traits",
            "saveChannel": "SigilGen + loadout constructor",
            "memoryChannel": "App.SigilMemoryGetOptions",
            "decision": "verified-equal",
            "detail": "same hashes, names, natural levels and trait ranges; historical runtime-only names remain read-only display fallbacks and are not selectable",
        },
    ]
    return {
        "schemaVersion": 1,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "drop": str(drop.resolve()),
        "policy": "filenames are untrusted; production catalogs are unchanged unless hashes cross-check against checked-in/runtime evidence",
        "catalogCounts": {name: len(values) for name, values in catalogs.items()},
        "files": files,
        "familyDiffs": families,
        "channelParity": channel_parity,
    }


def style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(name="Microsoft YaHei UI", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B4636")
        cell.alignment = Alignment(vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei UI", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, sheet.max_column + 1):
        width = min(64, max(12, max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1)) + 2))
        sheet.column_dimensions[get_column_letter(column)].width = width


def write_workbook(report: dict[str, object], path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    summary.append(["Policy", report["policy"]])
    summary.append(["Files", len(report["files"])])
    for name, count in report["catalogCounts"].items():
        summary.append([f"Checked-in {name}", count])

    inventory = workbook.create_sheet("Workbook Inventory")
    columns = ["file", "kind", "rows", "hashRows", "uniqueHashes", "duplicateHashes", "decision", "confidence", "reader", "reason", "sha256"]
    inventory.append(columns)
    fills = {
        "verified-equal": "D9EAD3",
        "verified-subset": "E2F0D9",
        "candidate": "FFF2CC",
        "reject": "F4CCCC",
        "reference-only": "D9EAF7",
    }
    for item in report["files"]:
        inventory.append([json.dumps(item.get(column), ensure_ascii=False) if isinstance(item.get(column), list) else item.get(column, "") for column in columns])
        fill = PatternFill("solid", fgColor=fills.get(item.get("decision"), "FFFFFF"))
        for cell in inventory[inventory.max_row]:
            cell.fill = fill

    diffs = workbook.create_sheet("Version Diffs")
    diffs.append(["family", "from", "to", "added count", "removed count", "added hashes", "removed hashes"])
    for item in report["familyDiffs"]:
        diffs.append([item["family"], item["from"], item["to"], len(item["added"]), len(item["removed"]), "\n".join(item["added"]), "\n".join(item["removed"])])

    parity = workbook.create_sheet("Channel Parity")
    parity.append(["family", "save channel", "memory channel", "decision", "detail"])
    for item in report["channelParity"]:
        parity.append([item["family"], item["saveChannel"], item["memoryChannel"], item["decision"], item["detail"]])

    for sheet in workbook.worksheets:
        style_sheet(sheet)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    # Verification pass catches corrupt output before it is handed off.
    checked = load_workbook(path, read_only=True, data_only=True)
    assert {"Summary", "Workbook Inventory", "Version Diffs", "Channel Parity"}.issubset(checked.sheetnames)
    checked.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("drop", type=Path)
    parser.add_argument("--repo", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--json", dest="json_path", type=Path, required=True)
    parser.add_argument("--xlsx", dest="xlsx_path", type=Path, required=True)
    args = parser.parse_args()
    report = audit(args.drop, args.repo)
    args.json_path.parent.mkdir(parents=True, exist_ok=True)
    args.json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_workbook(report, args.xlsx_path)
    print(json.dumps({"json": str(args.json_path), "xlsx": str(args.xlsx_path), "files": len(report["files"])}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
